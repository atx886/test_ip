from selenium import webdriver
import time
from datetime import timedelta, date

from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from post import *
from ph1 import *

import os

import json
import random
# import cehsi
import openpyxl
import requests
import time
from urllib.parse import unquote, quote
import re

from ph1 import *
from ph import *
from openpyxl import load_workbook, Workbook


def getphone():
    phone = getph()
    return phone


#
# def getid():
#     l = cehsi.outid()
#     name = l[0]
#     id = l[1]
#     return name, id


def gettext(r):
    t = r.text
    p = t[t.index("\\"):t.index("\",")].encode('ascii').decode('unicode_escape')
    print(p)
    return p


def getcode(t):
    code = getcd(t)
    return code


def gettoken(r):
    t = r.text
    p = t[t.index("bPfmSW"):t.index("\"}")].encode('ascii').decode('utf-8')
    print(p)
    return p


def getuserid(r):
    t = r.text
    p = re.findall(r"user_id\":(.+?),", t)
    # a = t.index("user_id")
    # print(a)
    # p = t[a+7:t.index("\"")]
    # p = t
    print(p[0])
    return p[0]


chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument('--headless')
chrome_options.add_argument('--no-sandbox')
chrome_options.add_argument('--disable-gpu')
chrome_options.add_argument('--disable-dev-shm-usage')
chromedriver = "/usr/bin/chromedriver"
os.environ["webdriver.chrome.driver"] = chromedriver


# d = webdriver.Chrome(chrome_options=chrome_options, executable_path=chromedriver)


def rw():
    time.sleep(1)


def writeexcle(phone):
    wb = load_workbook('a.xlsx')
    sheet = wb.active
    print(sheet)
    max_row = sheet.max_row + 1
    print(max_row)
    row_max = 'a' + str(max_row)
    print(row_max)
    sheet[row_max] = phone[0].strip('\'')
    wb.save('a.xlsx')


def zc(phone):
    d = webdriver.Chrome(chrome_options=chrome_options, executable_path=chromedriver)
    d.implicitly_wait(5)
    d.get('https://www.chaojijishi.com/h5/#/pages/login/register')
    d.find_element_by_xpath(
        '/html/body/uni-app/uni-page/uni-page-wrapper/uni-page-body/uni-view/uni-view[3]/uni-view[3]/uni-view[1]/uni-view/uni-input/div/input').send_keys(
        phone)
    rw()
    # 获取验证码
    t = d.find_element_by_xpath(
        '/html/body/uni-app/uni-page/uni-page-wrapper/uni-page-body/uni-view/uni-view[3]/uni-view[3]/uni-view[2]/uni-view/uni-text/span')
    d.execute_script("arguments[0].click();", t)
    rw()
    if t.text == "获取验证码":
        return
    rw()
    x = 0
    code = getcode(phone)
    while len(code) < 1:
        time.sleep(5)
        code = getcode(phone)
        x += 1
        print("尝试", x)
        if x == 6:
            closeph(phone)
            return
    # 验证码方式
    d.find_element_by_xpath(
        '/html/body/uni-app/uni-page/uni-page-wrapper/uni-page-body/uni-view/uni-view[3]/uni-view[3]/uni-view[2]/uni-input/div/input').send_keys(
        code)
    rw()
    t = d.find_element_by_xpath(
        '/html/body/uni-app/uni-page/uni-page-wrapper/uni-page-body/uni-view/uni-view[6]/uni-view/uni-view')
    d.execute_script("arguments[0].click();", t)
    rw()

    # 下一步
    t = d.find_element_by_xpath(
        '/html/body/uni-app/uni-page/uni-page-wrapper/uni-page-body/uni-view/uni-view[5]')
    d.execute_script("arguments[0].click();", t)
    rw()
    d.find_element_by_xpath(
        '/html/body/uni-app/uni-page/uni-page-wrapper/uni-page-body/uni-view/uni-view[4]/uni-view[3]/uni-view[1]/uni-input/div/input').send_keys(
        123456)
    rw()
    d.find_element_by_xpath(
        '/html/body/uni-app/uni-page/uni-page-wrapper/uni-page-body/uni-view/uni-view[4]/uni-view[3]/uni-view[2]/uni-input/div/input').send_keys(
        123456)
    rw()
    # 完成
    t = d.find_element_by_xpath(
        '/html/body/uni-app/uni-page/uni-page-wrapper/uni-page-body/uni-view/uni-view[5]')
    d.execute_script("arguments[0].click();", t)
    writeexcle(phone)
    d.close()
    # d.get('https://www.chaojijishi.com/h5/#/pages/subpack1/set/user-id-card-data?type=1')


i = 0
while i < 10:
    p = getphone()
    zc(p)
    i += 1
prw()
