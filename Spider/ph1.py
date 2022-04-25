from openpyxl import load_workbook, Workbook


def writeexcle():
    file = 'a.xlsx'
    wb = load_workbook(file)
    sheet = wb.active
    max_row = sheet.max_row
    row_max = 'a' + str(max_row)
    print(sheet)
    if sheet[row_max] is not None:
        a = sheet[row_max].value
        print(a)
        del sheet[row_max]
        wb.save(file)
        return a
    else:
        print("空了")
        return None



