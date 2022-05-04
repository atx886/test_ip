import re
from openpyxl import load_workbook, Workbook
import requests
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.header import Header
import time

token2 = 'dd03c30290d6-239153'
token1 = '425b1a2dc5b6-258930'
apiName = 'MY.431035769'


# 获取token
def gettk():
    url = 'http://api.miyun.pro/api/login?apiName=MY.431035769&password=qq200200'

    r = requests.get(url)

    print(r.content)


# 查询token是否失效
def checktk():
    url1 = 'http://api.miyun.pro/api/get_myinfo?token=' + token2
    r = requests.get(url1)
    print(r.content)


# 获取号码
def getph():
    url2 = 'http://api.miyun.pro/api/get_mobile?token=' + token2 + '&project_id=29756'
    r = requests.get(url2)
    print(r.content)
    t = r.text
    t = re.findall(r"mobile\":\"(.+?)\"", t)
    print('手机号是', t)
    return t


# 获取验证码
def getcd(t):
    url3 = 'http://api.miyun.pro/api/get_message?token=' + token2 + '&project_id=29756&phone_num=' + t[0]
    r = requests.get(url3)
    print(r.content)
    t = r.text
    t = re.findall(r"code\":\"(.+?)\",", t)
    print('验证码是', t)
    return t


# 拉黑号码
def closeph(t):
    url4 = 'http://api.miyun.pro/api/add_blacklist?token=' + token2 + '&project_id=29756&phone_num=' + t[0]
    r = requests.get(url4)
    print(r.content)


# 写入xcel
def writeexcle(phone):
    wb = load_workbook('a.xlsx')
    sheet = wb.active
    print(sheet)
    max_row = sheet.max_row + 1
    print(max_row)
    row_max = 'a' + str(max_row)
    print(row_max)
    sheet[row_max] = phone[0].strip('\'')
    wb.save('a.xlsx')


